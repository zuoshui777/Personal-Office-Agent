# Excel MCP

from pathlib import Path


def read_excel(path: str, sheet: str | None = None):
    file = Path(path)
    if not file.is_file():
        return {"error": "文件不存在"}

    from openpyxl import load_workbook

    wb = load_workbook(file, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])
    return {"rows": rows[:200]}


def analyze_excel(path: str):
    data = read_excel(path)
    if "error" in data:
        return data
    rows = data["rows"]
    return {
        "row_count": len(rows),
        "column_count": len(rows[0]) if rows else 0,
        "preview": rows[:10]
    }


def register(register_tool):
    register_tool("excel_read", "读取 Excel 文件", read_excel)
    register_tool("excel_analyze", "分析 Excel 文件", analyze_excel)
